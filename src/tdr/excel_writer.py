"""
Escritor de Excel para resultados TDR.

Genera Excel intermedio con 3 hojas:
1. Criterios RTM Profesionales
2. Experiencias Solicitadas
3. Información (metadatos)

Estilo profesional: headers azul oscuro, filas alternas, freeze panes, auto-filter.
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TDRExtraction

logger = logging.getLogger(__name__)

# ─── ESTILOS ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)

ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")   # azul claro filas pares
NO_FILL     = PatternFill("solid", fgColor="FFFFFF")

CELL_FONT      = Font(name="Arial", size=9)
CELL_FONT_BOLD = Font(name="Arial", size=9, bold=True)

THIN   = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
LEFT_C = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# ──────────────────────────────────────────────────────────────────────────────


class TDRExcelWriter:
    """Escritor de Excel para resultados TDR."""

    @staticmethod
    def write(extraction: TDRExtraction, output_path: Path) -> None:
        """Escribe extracción TDR a Excel."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"[Excel Writer] Generando Excel: {output_path}")

        wb = Workbook()
        wb.remove(wb.active)

        TDRExcelWriter._write_cargos_sheet(wb, extraction.cargos)
        TDRExcelWriter._write_experiencias_sheet(wb, extraction.experiencias)
        TDRExcelWriter._write_metadata_sheet(wb, extraction)

        wb.save(str(output_path))
        logger.info(f"[Excel Writer] Excel guardado en {output_path}")

    # ─── SHEET 1: CRITERIOS RTM PROFESIONALES ────────────────────────────────

    @staticmethod
    def _write_cargos_sheet(wb: Workbook, cargos: list) -> None:
        ws = wb.create_sheet("Criterios RTM Profesionales", 0)

        headers = [
            "Cargo y Profesión",
            "Años de Colegiado",
            "Requisito Mínimo\n(detallado + puntuación + máximo)",
            "Tipo Experiencia Similar",
            "Tiempo Adicional\n(Factores de Evaluación)",
            "Capacitación Solicitada",
        ]
        col_widths = [30, 15, 55, 25, 30, 30]

        # Headers
        _write_header_row(ws, headers)

        # Data
        for row_idx, cargo in enumerate(cargos, start=2):
            tipos_obra = ", ".join(cargo.tipos_obra_validos) if cargo.tipos_obra_validos else ""

            # Combinar requisito mínimo + puntuación en una celda
            requisito_completo = cargo.requisito_minimo_detallado or ""
            if cargo.puntuacion_experiencia:
                requisito_completo += f"\n\nPuntuación: {cargo.puntuacion_experiencia}"

            values = [
                f"{cargo.cargo}\n({cargo.profesion_requerida})",
                cargo.anos_minimos_colegiado,
                requisito_completo,
                tipos_obra,
                cargo.tiempo_adicional_evaluacion or "",
                cargo.capacitacion_solicitada or "",
            ]

            _write_data_row(ws, row_idx, values)

        # Formato final
        _apply_sheet_format(ws, col_widths, len(cargos))

        logger.info(f"[Excel Writer] Sheet 'Criterios RTM Profesionales': {len(cargos)} filas")

    # ─── SHEET 2: EXPERIENCIAS SOLICITADAS ───────────────────────────────────

    @staticmethod
    def _write_experiencias_sheet(wb: Workbook, experiencias: list) -> None:
        ws = wb.create_sheet("Experiencias Solicitadas", 1)

        headers = [
            "Tipo de Experiencia Válida",
            "Sector Válido",
            "Descripción Exacta\n(texto del documento)",
            "Página",
            "Experiencia Adicional\na Entregar",
            "Otros Factores\nde Evaluación",
        ]
        col_widths = [30, 20, 55, 10, 35, 30]

        # Headers
        _write_header_row(ws, headers)

        # Data
        for row_idx, exp in enumerate(experiencias, start=2):
            values = [
                exp.tipo_experiencia or "",
                exp.sector_valido or "",
                exp.descripcion_exacta or "",
                exp.pagina_documento if exp.pagina_documento else "",
                exp.experiencia_adicional or "",
                exp.otros_factores or "",
            ]
            _write_data_row(ws, row_idx, values)

        # Formato final
        _apply_sheet_format(ws, col_widths, len(experiencias))

        logger.info(f"[Excel Writer] Sheet 'Experiencias Solicitadas': {len(experiencias)} filas")

    # ─── SHEET 3: INFORMACIÓN / METADATOS ────────────────────────────────────

    @staticmethod
    def _write_metadata_sheet(wb: Workbook, extraction: TDRExtraction) -> None:
        ws = wb.create_sheet("Información", 2)

        # Sección: Datos de extracción
        meta_rows = [
            ["INFORMACIÓN DE EXTRACCIÓN", ""],
            ["PDF Procesado",       extraction.pdf_name],
            ["Total Páginas",       extraction.total_paginas],
            ["Cargos Extraídos",    len(extraction.cargos)],
            ["Experiencias Extraídas", len(extraction.experiencias)],
            ["Modelo LLM",         extraction.modelo_llm],
            ["Tiempo (s)",         f"{extraction.tiempo_procesamiento:.1f}"],
            ["Fecha Extracción",   extraction.fecha_extraccion.strftime("%Y-%m-%d %H:%M")],
            ["", ""],
        ]

        for row_data in meta_rows:
            ws.append(row_data)

        # Sección: Resumen de cargos
        row_offset = len(meta_rows) + 1
        ws.append(["RESUMEN DE CARGOS", "PROFESIÓN"])
        for cell in ws[row_offset]:
            cell.font = CELL_FONT_BOLD

        for cargo in extraction.cargos:
            ws.append([cargo.cargo, cargo.profesion_requerida])

        # Sección: Resumen de experiencias
        ws.append(["", ""])
        ws.append(["EXPERIENCIAS", "SECTOR"])

        for exp in extraction.experiencias:
            ws.append([exp.tipo_experiencia, exp.sector_valido])

        # Formato
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 40

        # Título en negrita
        ws["A1"].font = CELL_FONT_BOLD

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = cell.font or CELL_FONT

        logger.info("[Excel Writer] Sheet 'Información': metadatos agregados")


# ─── FUNCIONES AUXILIARES ─────────────────────────────────────────────────────

def _write_header_row(ws, headers: list) -> None:
    """Escribe fila de headers con estilo profesional."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    ws.row_dimensions[1].height = 35


def _write_data_row(ws, row_idx: int, values: list) -> None:
    """Escribe fila de datos con estilo alterno."""
    fill = ALT_FILL if row_idx % 2 == 0 else NO_FILL

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill   = fill
        cell.border = BORDER
        cell.font   = CELL_FONT

        # Columnas con texto largo → wrap + top-left
        if isinstance(value, str) and len(value) > 50:
            cell.alignment = LEFT
        else:
            cell.alignment = LEFT_C


def _apply_sheet_format(ws, col_widths: list, num_data_rows: int) -> None:
    """Aplica formato final: anchos, freeze panes, auto-filter, altura dinámica."""

    # Anchos de columna
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes: fija la fila de headers
    ws.freeze_panes = "A2"

    # Auto-filter en headers
    last_col = get_column_letter(len(col_widths))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # Altura dinámica según contenido
    for row_idx in range(2, num_data_rows + 2):
        max_lines = 1
        for col_idx in range(1, len(col_widths) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                col_width = col_widths[col_idx - 1]
                lineas = max(1, len(cell.value) // col_width + cell.value.count("\n"))
                max_lines = max(max_lines, lineas)

        ws.row_dimensions[row_idx].height = min(max(max_lines * 14, 20), 200)
